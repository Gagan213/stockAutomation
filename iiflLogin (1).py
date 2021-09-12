from XTSConnect import XTSConnect
import openpyxl,time
from radheUtils import conditionStopper


def login(API_KEY,API_SECRET,source,marketDataLogin=0):
    xt = XTSConnect(API_KEY, API_SECRET, source)
    if marketDataLogin==0:
        result=xt.interactive_login()
    elif marketDataLogin==1:
        result=xt.marketdata_login()
    if result.get('status')==1:
        return {'status':1,'blaze':xt,'result':result}
    else:
        return {'status':0,'msg':result.get('msg')}
    
    
def loginEasy(IIFLUserId,excelFile,marketDataLogin=0):
    wb1=openpyxl.load_workbook(excelFile)
    ws=wb1['IIFL Users']
    pointer=2
    while ws.cell(row=pointer,column=1).value!=None:
        if ws.cell(row=pointer,column=1).value==IIFLUserId:
            userId=ws.cell(row=pointer,column=1).value
            api=ws.cell(row=pointer,column=2).value
            apiSecret=ws.cell(row=pointer,column=3).value
            if marketDataLogin==1:
                api=ws.cell(row=pointer,column=4).value
                apiSecret=ws.cell(row=pointer,column=5).value
            loginResult=login(api,apiSecret,'WebAPI',marketDataLogin)
    
            return loginResult
        else:
            pointer+=1
    return {'status':0,'msg':'User Id Not Found in given File'}


def fetchOrderStatus(blaze,orderId):
    count=0
    while True:
        result=blaze.get_order_history(orderId)
        # print(f'Inside Fetch Order Status Function {orderId} {result}')
        if result.get('type')=='success':
            last=result.get('result',[])[-1]
            return {'status':last.get('OrderStatus'),'msg':last.get('CancelRejectReason')}
        elif result.get('data',{}).get('code')=='e-apirl-0004':
            time.sleep(1)
            continue
        elif result.get('code')=='e-orders-0001':
            if count<2:
                count+=1                
                time.sleep(1)
                continue
        print(result)
        return {'status':None,'msg':result.get('description')}

    # def checkstatus(orderId):    
    #     z=item.get('blaze').get_order_history(orderId)
    #     print(z)
    #     status=z.get('result')[-1].get('OrderStatus')
    #     return status   
    

def placeOrderGiveConfirmation(item):
    def true():
        pass
    blaze=item.get('kite')
    orderResult=blaze.place_order(item)
    # print(f'{orderResult} = order Result in place order confirmation')
    if orderResult.get('status'):
        def check():
            sts=fetchOrderStatus(blaze,orderResult.get('orderId'))
            return sts.get('status') in ['Filled','Rejected','Cancelled',None]
        conditionStopper(check,true,1)
        result=fetchOrderStatus(blaze,orderResult.get('orderId'))
        orderResult['confirm']=result.get('status')
        orderResult['msg']=result.get('msg')
    return orderResult




if __name__ == "__main__":
    x=loginEasy('IIFLJGDP','Users.xlsx',0)
    # print(x)

    
    # print('radhe radhe')
    # print(z)
    
     
    item={}
    item['exchangeSegment']='NSECM'
    item['exchangeInstrumentID']=2885
    item['productType']='NRML'
    item['orderType']='LIMIT'
    item['orderSide']='BUY'
    item['timeInForce']='DAY'
    item['disclosedQuantity']=0
    item['orderQuantity']=5
    item['limitPrice']=2
    item['stopPrice']=0
    item['orderUniqueIdentifier']='radhe'
    
    #
    item['blaze']=x.get('blaze')
    print(placeOrderGiveConfirmation(item))
    # def checkstatus(orderId):    
    #     z=item.get('blaze').get_order_history(orderId)
    #     print(z)
    #     status=z.get('result')[-1].get('OrderStatus')
    #     return status
    # print(*item)
    # z=item.get('blaze').place_order(item)
    # z=placeOrder(item)
    # print(z)
    # z=item.get('blaze').get_order_book()
    # for i in z.get('result'):
    #     status=fetchOrderStatus(item.get('blaze'),i.get('AppOrderID'))
    #     print(f'{i.get("AppOrderID")} = {status}')
    # print(fetchOrderStatus(item.get('blaze'),12))
    
# def placeOrder(item):
#     #Not NEcessary
#     print("Placing Order")
#     blaze=item.get('blaze')
#     exchangeSegment=item.get('exchangeSegment')
#     exchangeInstrumentID=item.get('exchangeInstrumentID')
#     productType=item.get('productType')
#     orderType=item.get('orderType')
#     orderSide=item.get('orderSide')
#     timeInForce=item.get('timeInForce')
#     disclosedQuantity=item.get('disclosedQuantity')
#     orderQuantity=item.get('orderQuantity')
#     limitPrice=item.get('limitPrice')
#     stopPrice=item.get('stopPrice')
#     orderUniqueIdentifier=item.get('orderUniqueIdentifier')
#     orderStatus=blaze.place_order(exchangeSegment=exchangeSegment,
#                     exchangeInstrumentID=exchangeInstrumentID,
#                     productType=productType,
#                     orderType=orderType,
#                     orderSide=orderSide,
#                     timeInForce=timeInForce,
#                     disclosedQuantity=disclosedQuantity,
#                     orderQuantity=orderQuantity,
#                     limitPrice=limitPrice,
#                     stopPrice=stopPrice,
#                     orderUniqueIdentifier=orderUniqueIdentifier)
#     print(orderStatus)
